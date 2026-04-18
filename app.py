"""
Экспертная система ИЛ - версия 2.0
Точка входа в приложение Streamlit.
"""

import streamlit as st
from st_keyup import st_keyup
from typing import List, Set
from datetime import datetime
import threading
import requests

from models import Laboratory
from data_loader import load_laboratories
from ui.sidebar import render_sidebar
from ui.tables import render_product_table, render_tnved_table, render_standard_table
from ui.results import render_intersection_result, render_indicators_result
from ui.styles import apply_styles

# ============================================================
# КОНФИГУРАЦИЯ ЛОГИРОВАНИЯ В GOOGLE SHEETS
# ============================================================
GOOGLE_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbw849AwsKvLlqi6e6tbvMwJ7IntMk6JQZ5RHJzAYb-L_idHzkyEFy_GMGd6PiHh-BYqpg/exec"


def write_log_async(lab, product, tnved, standards):
    """Отправляет лог в Google Sheets в фоновом потоке."""

    def _send():
        try:
            # Получаем IP-адрес
            ip = requests.get("https://api.ipify.org", timeout=3).text
        except:
            ip = "не определён"

        try:
            now = datetime.now()
            data = {
                "date": now.strftime("%d.%m.%Y"),
                "time": now.strftime("%H:%M:%S"),
                "lab": lab,
                "product": product,
                "tnved": tnved,
                "standards": standards,
                "ip": ip  # ← ДОБАВЛЕН IP
            }
            requests.post(GOOGLE_SCRIPT_URL, data=data, timeout=5)
        except Exception:
            pass

    threading.Thread(target=_send, daemon=True).start()


# ============================================================
# КОНФИГУРАЦИЯ СТРАНИЦЫ
# ============================================================
st.set_page_config(
    page_title="Экспертная система ИЛ v2",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded"
)

apply_styles()


# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================
def parse_input_text(text: str) -> list:
    """Парсит ввод пользователя в список."""
    if not text:
        return []
    for sep in [';', '\n', '\r', '\t']:
        text = text.replace(sep, ',')
    return [item.strip() for item in text.split(',') if item.strip()]


def clear_form():
    """Очищает поля формы и таблицы."""
    st.session_state.tnved_value = ""
    st.session_state.standard_value = ""
    st.session_state.indicator_value = ""  # ← ДОБАВИТЬ
    st.session_state.indicator_results = None  # ← ДОБАВИТЬ
    st.session_state.product_table = None
    st.session_state.tnved_table = None
    st.session_state.standard_table = None
    st.session_state.tables_built = False
    st.session_state.intersection_result = None
    st.session_state.intersection_ready = False
    st.session_state.indicators_result = None
    st.session_state.select_all_state = False
    st.session_state.export_ready = False
    for key in list(st.session_state.keys()):
        if key.startswith(('prod_cb_', 'tnved_cb_', 'std_cb_')):
            del st.session_state[key]
    st.session_state.product_checkboxes = {}
    st.session_state.tnved_checkboxes = {}
    st.session_state.standard_checkboxes = {}


def set_test_data():
    """Заполняет форму тестовыми данными (кроме продукции)."""
    st.session_state.tnved_value = "6115, 6116, 6117"
    st.session_state.standard_value = "МР 2915-82, МУ 4395-87"


# ============================================================
# ИНИЦИАЛИЗАЦИЯ СЕССИИ
# ============================================================
def init_session_state():
    """Инициализация session_state при первом запуске."""
    if 'labs_data' not in st.session_state:
        with st.spinner('Загрузка данных лабораторий...'):
            st.session_state.labs_data = load_laboratories()

        if st.session_state.labs_data:
            st.session_state.current_lab = list(st.session_state.labs_data.keys())[0]
            st.session_state.lab = st.session_state.labs_data[st.session_state.current_lab]
        else:
            st.error("Не удалось загрузить данные. Проверьте подключение к интернету.")
            st.stop()

    # Список всех показателей для автоподсказок (вычисляется один раз)
    if 'all_indicators' not in st.session_state:
        st.session_state.all_indicators = st.session_state.lab.get_all_indicators()


    if 'product_value' not in st.session_state:
        st.session_state.product_value = ''
    if 'tnved_value' not in st.session_state:
        st.session_state.tnved_value = ''
    if 'standard_value' not in st.session_state:
        st.session_state.standard_value = ''

    if 'product_table' not in st.session_state:
        st.session_state.product_table = None
    if 'tnved_table' not in st.session_state:
        st.session_state.tnved_table = None
    if 'standard_table' not in st.session_state:
        st.session_state.standard_table = None

    if 'selected_products' not in st.session_state:
        st.session_state.selected_products = set()
    if 'selected_tnved' not in st.session_state:
        st.session_state.selected_tnved = set()
    if 'selected_standards' not in st.session_state:
        st.session_state.selected_standards = set()

    if 'intersection_result' not in st.session_state:
        st.session_state.intersection_result = None
    if 'indicators_result' not in st.session_state:
        st.session_state.indicators_result = None

    if 'tables_built' not in st.session_state:
        st.session_state.tables_built = False
    if 'intersection_ready' not in st.session_state:
        st.session_state.intersection_ready = False

    if 'product_checkboxes' not in st.session_state:
        st.session_state.product_checkboxes = {}
    if 'tnved_checkboxes' not in st.session_state:
        st.session_state.tnved_checkboxes = {}
    if 'standard_checkboxes' not in st.session_state:
        st.session_state.standard_checkboxes = {}

    if 'search_mode' not in st.session_state:
        st.session_state.search_mode = "Поиск по стандартам"


init_session_state()

# ============================================================
# ЗАГОЛОВОК
# ============================================================
if st.session_state.current_lab == "ИЛ ЦТС":
    lab_class = "lab-cts"
    lab_icon = "🧬"
else:
    lab_class = "lab-tehexpert"
    lab_icon = "🧪"

st.markdown(f"""
<div style="margin-bottom: 20px;">
    <div class="main-title">🔬 Экспертная система ИЛ v2.0</div>
    <span class="lab-name {lab_class}">{lab_icon} {st.session_state.current_lab}</span>
</div>
""", unsafe_allow_html=True)

# ============================================================
# САЙДБАР
# ============================================================
render_sidebar(st.session_state.labs_data, st.session_state.current_lab)

# Убеждаемся, что search_mode точно есть
if 'search_mode' not in st.session_state:
    st.session_state.search_mode = "Поиск по стандартам"

# ============================================================
# ПОЛЕ ПРОДУКЦИИ (вне формы, с живыми подсказками)
# ============================================================
st.markdown("### 📦 Продукция")
product_input = st_keyup(
    "Введите название продукции",
    value=st.session_state.product_value,
    placeholder="например: белье, замша, стекло",
    key="product_keyup",
    label_visibility="collapsed"
)
st.session_state.product_value = product_input

if 'prev_product_value' not in st.session_state:
    st.session_state.prev_product_value = product_input

if product_input != st.session_state.prev_product_value:
    st.session_state.prev_product_value = product_input
    st.session_state.product_table = None
    st.session_state.tnved_table = None
    st.session_state.standard_table = None
    st.session_state.tables_built = False
    st.session_state.intersection_result = None
    st.session_state.intersection_ready = False
    st.session_state.indicators_result = None
    st.session_state.select_all_state = False
    st.session_state.export_ready = False
    for key in list(st.session_state.keys()):
        if key.startswith(('prod_cb_', 'tnved_cb_', 'std_cb_')):
            del st.session_state[key]
    st.session_state.product_checkboxes = {}
    st.session_state.tnved_checkboxes = {}
    st.session_state.standard_checkboxes = {}
    st.rerun()

if product_input and len(product_input) >= 2:
    lab = st.session_state.lab
    all_products = lab.get_all_products()
    suggestions = [p for p in all_products if product_input.lower() in p.lower()]
    if suggestions:
        st.markdown(f"**📋 Найдено в области ({len(suggestions)}):**")
        suggestions_text = ""
        for s in suggestions[:15]:
            suggestions_text += f"• {s}\n"
        if len(suggestions) > 15:
            suggestions_text += f"... и еще {len(suggestions) - 15}"
        st.text(suggestions_text)
    else:
        st.warning("❌ Ничего не найдено")

# ============================================================
# ПОЛЕ ПОКАЗАТЕЛЯ (вне формы, с живыми подсказками)
# ============================================================
if st.session_state.get("search_mode") == "Поиск по показателям":
    from ui.inputs import render_indicator_input

    indicator_input = render_indicator_input(st.session_state.get("indicator_value", ""))

    # Отслеживаем изменение показателя
    if 'prev_indicator_value' not in st.session_state:
        st.session_state.prev_indicator_value = indicator_input

    if indicator_input != st.session_state.prev_indicator_value:
        st.session_state.prev_indicator_value = indicator_input
        # Сбрасываем таблицы
        st.session_state.indicator_results = None
        st.session_state.tables_built = False

    st.session_state.indicator_value = indicator_input

# Живые подсказки для показателя (только в режиме "Показатели")
if st.session_state.get("search_mode") == "Поиск по показателям":
    indicator_value = st.session_state.get("indicator_value", "")
    if indicator_value and len(indicator_value) >= 2:
        all_indicators = st.session_state.get("all_indicators", [])
        suggestions = [p for p in all_indicators if indicator_value.lower() in p.lower()]
        if suggestions:
            st.markdown(f"**📋 Найдено показателей ({len(suggestions)}):**")
            suggestions_text = ""
            for s in suggestions[:15]:
                suggestions_text += f"• {s}\n"
            if len(suggestions) > 15:
                suggestions_text += f"... и еще {len(suggestions)-15}"
            st.text(suggestions_text)
        else:
            st.warning("❌ Показатель не найден в области")


# ============================================================
# ОТСЛЕЖИВАНИЕ ИЗМЕНЕНИЙ ТН ВЭД И СТАНДАРТОВ
# ============================================================
if 'prev_tnved_value' not in st.session_state:
    st.session_state.prev_tnved_value = st.session_state.tnved_value
if 'prev_standard_value' not in st.session_state:
    st.session_state.prev_standard_value = st.session_state.standard_value

# ============================================================
# ФОРМА (ТН ВЭД, Стандарты, кнопки)
# ============================================================
with st.form(key="input_form"):
    col1, col2 = st.columns(2)
    with col1:
        tnved_input = st.text_area(
            "📟 Коды ТН ВЭД",
            value=st.session_state.tnved_value,
            placeholder="6115, 6116, 6117",
            height=80
        )
    with col2:
        if st.session_state.get("search_mode") == "Поиск по показателям":
            # Неактивное поле для сохранения вида
            st.text_area(
                "📄 Стандарты",
                value="",
                placeholder="(не используется в этом режиме)",
                height=80,
                disabled=True,
                key="standards_disabled"
            )
        else:
            standard_input = st.text_area(
                "📄 Стандарты",
                value=st.session_state.standard_value,
                placeholder="ГОСТ 1234-2020, МР 2915-82",
                height=80
            )

    col_btn1, col_btn2, col_btn3 = st.columns(3)
    with col_btn1:
        build_btn = st.form_submit_button("📊 Построить таблицы", type="primary", use_container_width=True)
    with col_btn2:
        clear_btn = st.form_submit_button("🗑️ Очистить", on_click=clear_form, use_container_width=True)
    with col_btn3:
        test_btn = st.form_submit_button("⚡ Тест", on_click=set_test_data, use_container_width=True)

if build_btn or clear_btn or test_btn:
    # Проверяем, изменились ли ТН ВЭД или Стандарты
    tnved_changed = tnved_input != st.session_state.prev_tnved_value
    standard_changed = False

    # standard_input существует только в режиме "Стандарты"
    if st.session_state.get("search_mode") != "Поиск по показателям":
        standard_changed = standard_input != st.session_state.prev_standard_value

    if tnved_changed or standard_changed:
        st.session_state.product_table = None
        st.session_state.tnved_table = None
        st.session_state.standard_table = None
        st.session_state.tables_built = False
        st.session_state.intersection_result = None
        st.session_state.intersection_ready = False
        st.session_state.indicators_result = None
        st.session_state.indicator_results = None
        st.session_state.select_all_state = False
        st.session_state.export_ready = False
        for key in list(st.session_state.keys()):
            if key.startswith(('prod_cb_', 'tnved_cb_', 'std_cb_')):
                del st.session_state[key]
        st.session_state.product_checkboxes = {}
        st.session_state.tnved_checkboxes = {}
        st.session_state.standard_checkboxes = {}

    st.session_state.tnved_value = tnved_input
    st.session_state.prev_tnved_value = tnved_input

    # Сохраняем standard_value только в режиме "Стандарты"
    if st.session_state.get("search_mode") != "Поиск по показателям":
        st.session_state.standard_value = standard_input
        st.session_state.prev_standard_value = standard_input
# ============================================================
# ОБРАБОТКА НАЖАТИЯ "ПОСТРОИТЬ ТАБЛИЦЫ"
# ============================================================
if build_btn and st.session_state.product_value:

    write_log_async(
        lab=st.session_state.current_lab,
        product=st.session_state.product_value,
        tnved=st.session_state.tnved_value,
        standards=st.session_state.standard_value
    )

    # РЕЖИМ "ПОКАЗАТЕЛИ"
    if st.session_state.get("search_mode") == "Поиск по показателям":
        from logic import find_standards_by_indicator

        lab = st.session_state.lab
        product_query = st.session_state.product_value
        tnved_list = parse_input_text(st.session_state.tnved_value)
        indicator_query = st.session_state.get("indicator_value", "")

        if not indicator_query:
            st.warning("❌ Введите показатель")
        else:
            result = find_standards_by_indicator(lab, product_query, tnved_list, indicator_query)

            # Проверяем, есть ли хоть один результат
            if not result or all(len(rows) == 0 for rows in result.values()):
                st.session_state.indicator_results = None
                st.session_state.indicator_warning = "❌ Стандарты с указанным показателем не найдены для данной продукции и ТН ВЭД"
            else:
                st.session_state.indicator_results = result
                st.session_state.indicator_warning = None

            st.session_state.tables_built = True
            st.rerun()

    # РЕЖИМ "СТАНДАРТЫ"
    else:
        from logic import search_products, filter_by_tnved, filter_by_standard_and_tnved, group_sections_for_display

        lab = st.session_state.lab
        product_query = st.session_state.product_value

        products = search_products(lab, product_query)
        if products:
            rows = []
            for product_name, sections in products.items():
                rows.append({
                    'Формулировка продукции': product_name,
                    'Разделы': group_sections_for_display(sections),
                    '_sections': sections
                })
            st.session_state.product_table = rows
        else:
            st.warning("❌ Продукция не найдена")
            st.session_state.product_table = None

        tnved_list = parse_input_text(st.session_state.tnved_value)
        if tnved_list and st.session_state.product_table:
            rows = []
            all_product_sections = set().union(*[row['_sections'] for row in st.session_state.product_table])

            for code in tnved_list:
                sections = set()
                matched_codes = set()

                for section in lab.sections:
                    for tn in section.tnved_codes:
                        if code.startswith(tn):
                            if section.full_id in all_product_sections:
                                sections.add(section.full_id)
                                matched_codes.add(tn)

                display_code = list(matched_codes)[0] if matched_codes else code

                rows.append({
                    'ТН ВЭД': display_code,
                    'Разделы': group_sections_for_display(sections) if sections else '❌ Не найден',
                    '_sections': sections
                })
            st.session_state.tnved_table = rows
        else:
            st.session_state.tnved_table = None

        standard_list = parse_input_text(st.session_state.standard_value)
        st.session_state.standard_list = standard_list

        if standard_list:
            rows = []
            search_tnved = tnved_list if tnved_list else [None]

            for std in standard_list:
                std_lower = std.lower()
                std_found = False

                for code in search_tnved:
                    for section in lab.sections:
                        full_std_name = std
                        found_in_section = False
                        for s in section.standards:
                            if std_lower in s.lower():
                                full_std_name = s
                                found_in_section = True
                                std_found = True
                                break

                        if not found_in_section:
                            continue

                        if code is not None:
                            tnved_found = False
                            for tn in section.tnved_codes:
                                if code.startswith(tn):
                                    tnved_found = True
                                    matched_tn = tn
                                    break
                            if not tnved_found:
                                continue
                        else:
                            matched_tn = '—'

                        rows.append({
                            'Стандарт': full_std_name,
                            'ТН ВЭД': matched_tn,
                            'Разделы': group_sections_for_display({section.full_id}),
                            '_sections': {section.full_id}
                        })

                if not std_found:
                    for code in search_tnved:
                        rows.append({
                            'Стандарт': std,
                            'ТН ВЭД': code if code else '—',
                            'Разделы': '❌ Стандарт не найден',
                            '_sections': set()
                        })

            st.session_state.standard_table = rows if rows else None
        else:
            st.session_state.standard_table = None

        st.session_state.tables_built = True
        st.session_state.intersection_result = None
        st.session_state.intersection_ready = False
        st.session_state.indicators_result = None
        st.session_state.select_all_state = False

        for key in list(st.session_state.keys()):
            if key.startswith(('prod_cb_', 'tnved_cb_', 'std_cb_')):
                del st.session_state[key]
        st.session_state.product_checkboxes = {}
        st.session_state.tnved_checkboxes = {}
        st.session_state.standard_checkboxes = {}
        st.session_state.selected_products = set()
        st.session_state.selected_tnved = set()
        st.session_state.selected_standards = set()

        st.rerun()

# ============================================================
# ПАНЕЛЬ КНОПОК УПРАВЛЕНИЯ
# ============================================================

tables_built = st.session_state.get('tables_built', False)
intersection_ready = st.session_state.get('intersection_ready', False)
indicators_ready = st.session_state.get('indicators_result') is not None

st.markdown("---")

# Первая строка кнопок (только для режима "Стандарты")
if st.session_state.get("search_mode") != "Поиск по показателям":
    col1, col2, col3 = st.columns(3)

    with col1:
        if 'select_all_state' not in st.session_state:
            st.session_state.select_all_state = False

        button_label = "🌍 Убрать все" if st.session_state.select_all_state else "🌍 Выбрать все"
        button_type = "secondary" if st.session_state.select_all_state else "primary"

        if st.button(button_label, type=button_type, disabled=not tables_built, use_container_width=True,
                     key="btn_select_all"):
            new_value = not st.session_state.select_all_state
            if st.session_state.get('product_table'):
                for i in range(len(st.session_state.product_table)):
                    st.session_state[f"prod_cb_{i}"] = new_value
                    st.session_state.product_checkboxes[i] = new_value
            if st.session_state.get('tnved_table'):
                for i in range(len(st.session_state.tnved_table)):
                    st.session_state[f"tnved_cb_{i}"] = new_value
                    st.session_state.tnved_checkboxes[i] = new_value
            if st.session_state.get('standard_table'):
                for i in range(len(st.session_state.standard_table)):
                    st.session_state[f"std_cb_{i}"] = new_value
                    st.session_state.standard_checkboxes[i] = new_value
            st.session_state.select_all_state = new_value
            st.rerun()

    with col2:
        if st.button("🔍 Пересечение", type="primary", disabled=not tables_built, use_container_width=True,
                     key="btn_intersect"):
            selected_product_sections = set()
            selected_tnved_sections = set()
            selected_standard_sections = set()

            if st.session_state.get('product_table'):
                for idx, row in enumerate(st.session_state.product_table):
                    if st.session_state.get(f"prod_cb_{idx}", False):
                        selected_product_sections.update(row['_sections'])

            if st.session_state.get('tnved_table'):
                for idx, row in enumerate(st.session_state.tnved_table):
                    if st.session_state.get(f"tnved_cb_{idx}", False):
                        selected_tnved_sections.update(row['_sections'])

            if st.session_state.get('standard_table'):
                for idx, row in enumerate(st.session_state.standard_table):
                    if st.session_state.get(f"std_cb_{idx}", False):
                        selected_standard_sections.update(row['_sections'])

            if selected_tnved_sections:
                common = selected_product_sections & selected_tnved_sections
                if selected_standard_sections:
                    common = common & selected_standard_sections
            else:
                common = selected_product_sections

            st.session_state.intersection_result = common
            st.session_state.intersection_ready = bool(common)
            st.rerun()

    with col3:
        if st.button("📋 Показатели", disabled=not intersection_ready, use_container_width=True, key="btn_indicators"):
            from logic import extract_indicators

            lab = st.session_state.lab
            common_sections = st.session_state.intersection_result

            selected_std_names = []
            if st.session_state.get('standard_table'):
                for idx, row in enumerate(st.session_state.standard_table):
                    if st.session_state.get(f"std_cb_{idx}", False):
                        selected_std_names.append(row['Стандарт'])

            indicators = extract_indicators(lab, common_sections, selected_std_names)
            st.session_state.indicators_result = indicators
            st.session_state.export_ready = False
            st.rerun()

# ============================================================
# ЭКСПОРТ В EXCEL
# ============================================================
col_exp1, col_exp2, col_exp3, col_exp4 = st.columns([1, 1, 1, 1])

with col_exp2:
    # Определяем, активна ли кнопка
    if st.session_state.get("search_mode") == "Поиск по показателям":
        export_disabled = not st.session_state.get("indicator_results")
    else:
        export_disabled = not st.session_state.get("indicators_result")

    if st.button("📄 Подготовить отчёт",
                 disabled=export_disabled,
                 use_container_width=True,
                 key="btn_prepare_export"):

        # ============================================================
        # ЭКСПОРТ ДЛЯ РЕЖИМА "ПОКАЗАТЕЛИ"
        # ============================================================
        if st.session_state.get("search_mode") == "Поиск по показателям" and st.session_state.get("indicator_results"):
            import pandas as pd
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчёт"

            current_row = 1

            border_side = Side(style='thin')
            border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)


            def add_section_title(title, row):
                ws.merge_cells(f'A{row}:D{row}')
                cell = ws[f'A{row}']
                cell.value = title
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                return row + 1


            def add_dataframe_to_sheet(df, start_row):
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        if r_idx == start_row:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                return start_row + len(df) + 2


            # --- ПАРАМЕТРЫ ПОИСКА ---
            current_row = add_section_title("🔍 ПАРАМЕТРЫ ПОИСКА", current_row)
            params_data = [
                ['Лаборатория', st.session_state.current_lab],
                ['Продукция', st.session_state.product_value],
                ['Коды ТН ВЭД', st.session_state.tnved_value],
                ['Показатель', st.session_state.get("indicator_value", "")]
            ]
            df_params = pd.DataFrame(params_data, columns=['Параметр', 'Значение'])
            current_row = add_dataframe_to_sheet(df_params, current_row)

            # --- ТАБЛИЦА 1. ПРОДУКЦИЯ ---
            if st.session_state.get('product_table'):
                current_row = add_section_title("📦 ТАБЛИЦА 1. ПРОДУКЦИЯ", current_row)
                df1 = pd.DataFrame(st.session_state.product_table)
                if '_sections' in df1.columns:
                    df1 = df1.drop(columns=['_sections'])
                current_row = add_dataframe_to_sheet(df1, current_row)

            # --- ТАБЛИЦА 2. ТН ВЭД ---
            if st.session_state.get('tnved_table'):
                current_row = add_section_title("📟 ТАБЛИЦА 2. ТН ВЭД", current_row)
                df2 = pd.DataFrame(st.session_state.tnved_table)
                if '_sections' in df2.columns:
                    df2 = df2.drop(columns=['_sections'])
                current_row = add_dataframe_to_sheet(df2, current_row)

            # --- РЕЗУЛЬТАТЫ ПОИСКА ПО ПОКАЗАТЕЛЮ ---
            if st.session_state.get('indicator_results'):
                current_row = add_section_title("📋 РЕЗУЛЬТАТЫ ПОИСКА ПО ПОКАЗАТЕЛЮ", current_row)

                for tnved_code, rows in st.session_state.indicator_results.items():
                    if tnved_code == "без ТН ВЭД":
                        section_title = "🔍 Без ТН ВЭД"
                    else:
                        section_title = f"📟 ТН ВЭД: {tnved_code}"

                    ws.merge_cells(f'A{current_row}:D{current_row}')
                    cell = ws[f'A{current_row}']
                    cell.value = section_title
                    cell.font = Font(bold=True, size=11)
                    cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                    cell.alignment = Alignment(horizontal='left')
                    current_row += 1

                    if rows:
                        # Копируем данные, чтобы не испортить интерфейс
                        excel_rows = []
                        for row in rows:
                            excel_row = row.copy()
                            if 'Значение' in excel_row:
                                excel_row['Значение'] = excel_row['Значение'].replace('<br>', '\n').replace('\r\n',
                                                                                                            '\n')
                            excel_rows.append(excel_row)

                        df_ind = pd.DataFrame(excel_rows)
                        if 'ТН ВЭД' in df_ind.columns:
                            df_ind = df_ind.drop(columns=['ТН ВЭД'])
                        current_row = add_dataframe_to_sheet(df_ind, current_row)
                    else:
                        ws.merge_cells(f'A{current_row}:D{current_row}')
                        cell = ws[f'A{current_row}']
                        cell.value = "❌ Стандарты не найдены"
                        cell.alignment = Alignment(horizontal='left')
                        current_row += 2

                    current_row += 1

            # Автоподбор ширины столбцов
            for col_letter in ['A', 'B', 'C', 'D']:
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    cell = ws[f'{col_letter}{row}']
                    try:
                        if cell.value:
                            lines = str(cell.value).split('\n')
                            for line in lines:
                                max_length = max(max_length, len(line))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 60)
                ws.column_dimensions[col_letter].width = adjusted_width

            wb.save(output)
            output.seek(0)
            st.session_state.export_data = output.getvalue()
            st.session_state.export_ready = True
            st.rerun()

        # ============================================================
        # ЭКСПОРТ ДЛЯ РЕЖИМА "СТАНДАРТЫ"
        # ============================================================
        elif st.session_state.indicators_result:
            import pandas as pd
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows

            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Отчёт"

            current_row = 1

            border_side = Side(style='thin')
            border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
            highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")


            def add_section_title(title, row):
                ws.merge_cells(f'A{row}:B{row}')
                cell = ws[f'A{row}']
                cell.value = title
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                return row + 1


            def add_dataframe_to_sheet(df, start_row):
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        if r_idx == start_row:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                return start_row + len(df) + 2


            current_row = add_section_title("🔍 ПАРАМЕТРЫ ПОИСКА", current_row)
            params_data = [
                ['Лаборатория', st.session_state.current_lab],
                ['Продукция', st.session_state.product_value],
                ['Коды ТН ВЭД', st.session_state.tnved_value],
                ['Стандарты', st.session_state.standard_value]
            ]
            df_params = pd.DataFrame(params_data, columns=['Параметр', 'Значение'])
            current_row = add_dataframe_to_sheet(df_params, current_row)

            if st.session_state.get('product_table'):
                current_row = add_section_title("📦 ТАБЛИЦА 1. ПРОДУКЦИЯ", current_row)
                df1 = pd.DataFrame(st.session_state.product_table)
                if '_sections' in df1.columns:
                    df1 = df1.drop(columns=['_sections'])
                current_row = add_dataframe_to_sheet(df1, current_row)

            if st.session_state.get('tnved_table'):
                current_row = add_section_title("📟 ТАБЛИЦА 2. ТН ВЭД", current_row)
                df2 = pd.DataFrame(st.session_state.tnved_table)
                if '_sections' in df2.columns:
                    df2 = df2.drop(columns=['_sections'])
                current_row = add_dataframe_to_sheet(df2, current_row)

            if st.session_state.get('standard_table'):
                current_row = add_section_title("📄 ТАБЛИЦА 3. СТАНДАРТЫ", current_row)
                df3 = pd.DataFrame(st.session_state.standard_table)
                if '_sections' in df3.columns:
                    df3 = df3.drop(columns=['_sections'])
                if 'Стандарт' in df3.columns and 'ТН ВЭД' in df3.columns:
                    df3['Стандарт (ТН ВЭД)'] = df3['Стандарт'] + ' (' + df3['ТН ВЭД'] + ')'
                    df3 = df3[['Стандарт (ТН ВЭД)', 'Разделы']]
                current_row = add_dataframe_to_sheet(df3, current_row)

            if st.session_state.get('intersection_result'):
                current_row = add_section_title("✅ РЕЗУЛЬТАТ ПЕРЕСЕЧЕНИЯ", current_row)
                from logic import group_sections_for_display

                intersection_str = group_sections_for_display(st.session_state.intersection_result)
                ws.merge_cells(f'A{current_row}:B{current_row}')
                for col in ['A', 'B']:
                    cell = ws[f'{col}{current_row}']
                    cell.border = border
                cell = ws[f'A{current_row}']
                cell.value = f"Общие разделы: {intersection_str}"
                cell.font = Font(bold=True)
                cell.fill = highlight_fill
                cell.alignment = Alignment(wrap_text=True)
                current_row += 2

            if st.session_state.indicators_result:
                current_row = add_section_title("📋 ПОКАЗАТЕЛИ ИСПЫТАНИЙ", current_row)

                from collections import defaultdict

                grouped = defaultdict(lambda: defaultdict(list))

                for ind in st.session_state.indicators_result:
                    std = ind.get('standard', '—')
                    sec = ind.get('section', '—')
                    name = ind.get('name', '')
                    range_val = ind.get('range', '').replace('<br>', '\n').replace('\r\n', '\n').strip()

                    if '_' in sec:
                        src, num = sec.split('_', 1)
                        display_sec = f"{src}:{num}"
                    else:
                        display_sec = sec

                    grouped[std][display_sec].append({
                        'Показатель': name,
                        'Значение': range_val
                    })

                for std, sections in grouped.items():
                    ws.merge_cells(f'A{current_row}:B{current_row}')
                    cell = ws[f'A{current_row}']
                    cell.value = std
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                    cell.alignment = Alignment(horizontal='left')
                    current_row += 1

                    for sec, indicators in sections.items():
                        ws.merge_cells(f'A{current_row}:B{current_row}')
                        cell = ws[f'A{current_row}']
                        cell.value = sec
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                        cell.alignment = Alignment(horizontal='left')
                        current_row += 1

                        ind_data = []
                        for ind in indicators:
                            ind_data.append({
                                'Показатель': ind['Показатель'],
                                'Значение': ind['Значение']
                            })

                        df_ind = pd.DataFrame(ind_data)
                        current_row = add_dataframe_to_sheet(df_ind, current_row)

                    current_row += 1

            for col_letter in ['A', 'B', 'C', 'D']:
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    cell = ws[f'{col_letter}{row}']
                    try:
                        if cell.value:
                            lines = str(cell.value).split('\n')
                            for line in lines:
                                max_length = max(max_length, len(line))
                    except:
                        pass
                adjusted_width = min(max_length + 3, 60)
                ws.column_dimensions[col_letter].width = adjusted_width

            wb.save(output)
            output.seek(0)
            st.session_state.export_data = output.getvalue()
            st.session_state.export_ready = True
            st.rerun()
with col_exp3:
    if st.session_state.get('export_ready', False):
        st.download_button(
            label="📥 Скачать",
            data=st.session_state.export_data,
            file_name=f"Экспертная_оценка_{st.session_state.current_lab}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_export_file",
            use_container_width=True
        )
    else:
        st.button("📥 Скачать", disabled=True, use_container_width=True, key="btn_download_disabled")

# ============================================================
# ОТОБРАЖЕНИЕ РЕЗУЛЬТАТОВ
# ============================================================
if st.session_state.intersection_result is not None:
    render_intersection_result(st.session_state.intersection_result)

if st.session_state.indicators_result is not None:
    render_indicators_result(st.session_state.indicators_result)

if st.session_state.get('standard_warning_msg'):
    st.warning(st.session_state.standard_warning_msg)
    del st.session_state.standard_warning_msg

# ============================================================
# ОТОБРАЖЕНИЕ ПРЕДУПРЕЖДЕНИЯ ДЛЯ РЕЖИМА "ПОКАЗАТЕЛИ"
# ============================================================
if st.session_state.get("indicator_warning"):
    st.warning(st.session_state.indicator_warning)
    # Очищаем, чтобы не показывать повторно
    st.session_state.indicator_warning = None



# ============================================================
# ОТОБРАЖЕНИЕ РЕЗУЛЬТАТОВ ДЛЯ РЕЖИМА "ПОКАЗАТЕЛИ"
# ============================================================
if st.session_state.get("search_mode") == "Поиск по показателям" and st.session_state.get("indicator_results"):
    st.markdown("---")
    st.subheader("📋 Результаты поиска по показателю")

    results = st.session_state.indicator_results

    if not results:
        st.warning("❌ Стандарты с указанным показателем не найдены для данной продукции и ТН ВЭД")
    else:
        for tnved_code, rows in results.items():
            if tnved_code == "без ТН ВЭД":
                st.markdown("### 🔍 Поиск без ТН ВЭД")
            else:
                st.markdown(f"### 📟 ТН ВЭД: {tnved_code}")

            if rows:
                import pandas as pd

                df = pd.DataFrame(rows)
                if 'ТН ВЭД' in df.columns:
                    df = df.drop(columns=['ТН ВЭД'])
                df = df[['Стандарт', 'Раздел', 'Показатель', 'Значение']]

                # Рендерим как HTML-таблицу (лучше контролируется)
                html_table = df.to_html(index=False, classes='indicator-results-table', escape=False)
                st.markdown(html_table, unsafe_allow_html=True)
            else:
                st.info("ℹ️ Для этого ТН ВЭД стандарты не найдены")
            st.markdown("---")


# ============================================================
# ОТОБРАЖЕНИЕ ТАБЛИЦ
# ============================================================
if st.session_state.tables_built:
    st.markdown("---")

    if 'tables_just_built' not in st.session_state:
        st.session_state.tables_just_built = False

    if not st.session_state.tables_just_built:
        for key in list(st.session_state.keys()):
            if key.startswith(('prod_cb_', 'tnved_cb_', 'std_cb_')):
                st.session_state[key] = False
        st.session_state.product_checkboxes = {}
        st.session_state.tnved_checkboxes = {}
        st.session_state.standard_checkboxes = {}
        st.session_state.selected_products = set()
        st.session_state.selected_tnved = set()
        st.session_state.selected_standards = set()
        st.session_state.select_all_state = False
        st.session_state.tables_just_built = True

    if st.session_state.product_table:
        selected = render_product_table(
            st.session_state.product_table,
            st.session_state.selected_products
        )
        st.session_state.selected_products = selected

    if st.session_state.tnved_table:
        selected = render_tnved_table(
            st.session_state.tnved_table,
            st.session_state.selected_tnved
        )
        st.session_state.selected_tnved = selected

    if st.session_state.standard_table:
        selected = render_standard_table(
            st.session_state.standard_table,
            st.session_state.selected_standards
        )
        st.session_state.selected_standards = selected